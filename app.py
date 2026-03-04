import streamlit as st
import requests
import math
import io
from itertools import combinations
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Distâncias entre Cidades", page_icon="🗺️", layout="wide")

# ── Índice de Serpentividade ───────────────────────────────────────────────────
# < 1.20  → Direta     (verde)
# 1.20–1.50 → Moderada (amarelo)
# > 1.50  → Sinuosa    (vermelho)
def serp_class(idx):
    if idx is None:   return None, None, None
    if idx < 1.20:    return "Direta",   "#15803d", "#dcfce7"
    if idx < 1.50:    return "Moderada", "#92400e", "#fef3c7"
    return             "Sinuosa",  "#991b1b", "#fee2e2"

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.header-box {
    background: linear-gradient(135deg, #0f1117 60%, #1a2744);
    color: #f5f0e8; padding: 28px 36px; border-radius: 8px; margin-bottom: 24px;
}
.header-tag { font-family:'DM Mono',monospace; font-size:.7rem; letter-spacing:3px;
    text-transform:uppercase; color:#c84b2f; margin-bottom:8px; }
.header-title { font-size:2rem; font-weight:300; line-height:1.2; margin:0; }
.header-sub { font-family:'DM Mono',monospace; font-size:.75rem; color:#8a8070;
    margin-top:6px; letter-spacing:.5px; }
.city-chip { display:inline-flex; align-items:center; gap:6px; background:#eff6ff;
    border:1.5px solid #bfdbfe; border-radius:20px; padding:4px 12px; font-size:.82rem;
    color:#1e40af; font-weight:600; margin:3px; }
.chip-coords { font-family:'DM Mono',monospace; font-size:.65rem; color:#64748b; }
/* Tabela principal */
.dist-table { width:100%; border-collapse:collapse; font-size:.8rem; margin-top:8px; }
.dist-table thead th { background:#0f1117; color:#f5f0e8; padding:9px 10px;
    text-align:center; font-weight:600; white-space:nowrap; font-size:.72rem; }
.dist-table thead th:first-child { background:#1a1d26; text-align:left; min-width:140px;
    font-family:'DM Mono',monospace; font-size:.63rem; letter-spacing:1px; text-transform:uppercase; }
.dist-table tbody tr:nth-child(even) { background:#f8fafc; }
.dist-table tbody tr:hover td { background:#f0f9ff !important; }
.dist-table tbody td { padding:8px 10px; text-align:center; border-bottom:1px solid #e2e8f0;
    border-right:1px solid #f1f5f9; vertical-align:middle; }
.dist-table tbody td:first-child { font-weight:600; text-align:left; color:#0f1117;
    white-space:nowrap; background:#fafaf8; border-right:2px solid #cbd5e1; }
.cell-self { background:#f1f5f9 !important; color:#94a3b8; }
.val-line  { color:#1e40af; font-weight:700; font-size:.82rem; }
.val-road  { color:#166534; font-weight:700; font-size:.82rem; }
.val-label { font-size:.6rem; color:#94a3b8; font-family:'DM Mono',monospace; }
.val-nd    { color:#fca5a5; font-size:.7rem; }
.serp-badge { display:inline-block; padding:1px 7px; border-radius:10px;
    font-family:'DM Mono',monospace; font-size:.68rem; font-weight:700; }
/* Tabela serpentividade */
.serp-table { width:100%; border-collapse:collapse; font-size:.82rem; margin-top:8px; }
.serp-table thead th { background:#0f1117; color:#f5f0e8; padding:9px 12px;
    text-align:center; font-weight:600; font-size:.73rem; }
.serp-table thead th:first-child,.serp-table thead th:nth-child(2) { text-align:left; background:#1a1d26; }
.serp-table tbody tr:nth-child(even) { background:#f8fafc; }
.serp-table tbody tr:hover td { background:#f0f9ff !important; }
.serp-table tbody td { padding:8px 12px; border-bottom:1px solid #e2e8f0; vertical-align:middle; }
.serp-table tbody td:nth-child(3),.serp-table tbody td:nth-child(4),.serp-table tbody td:nth-child(5) { text-align:center; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="header-box">
    <div class="header-tag">// análise geoespacial · Brasil</div>
    <h1 class="header-title">🗺️ Distâncias entre <em>Cidades</em></h1>
    <div class="header-sub">linha reta · por estrada · índice de serpentividade · sedes municipais</div>
</div>
""", unsafe_allow_html=True)

# ── Dados ──────────────────────────────────────────────────────────────────────
CAPITALS_DEFAULT = [
    # ── Capitais ──────────────────────────────────────────────────────────────
    {"name":"Rio Branco",          "state":"AC","lat":-9.97499, "lon":-67.82471,"address":"R. Benjamim Constant, 945","tipo":"Capital"},
    {"name":"Maceió",              "state":"AL","lat":-9.66583, "lon":-35.73528,"address":"Praça Thomaz Espíndola, s/n","tipo":"Capital"},
    {"name":"Macapá",              "state":"AP","lat": 0.03444, "lon":-51.06639,"address":"Av. Iracema Carvão do Nascimento, 600","tipo":"Capital"},
    {"name":"Manaus",              "state":"AM","lat":-3.10194, "lon":-60.02500,"address":"Av. Brasil, 2971 - Compensa","tipo":"Capital"},
    {"name":"Salvador",            "state":"BA","lat":-12.97111,"lon":-38.51083,"address":"Praça Municipal, s/n","tipo":"Capital"},
    {"name":"Fortaleza",           "state":"CE","lat":-3.72389, "lon":-38.54306,"address":"Av. Desembargador Moreira, 310","tipo":"Capital"},
    {"name":"Brasília",            "state":"DF","lat":-15.78361,"lon":-47.89833,"address":"Palácio do Buriti","tipo":"Capital"},
    {"name":"Vitória",             "state":"ES","lat":-20.31944,"lon":-40.33778,"address":"Av. Marechal Mascarenhas de Moraes, 1927","tipo":"Capital"},
    {"name":"Goiânia",             "state":"GO","lat":-16.67861,"lon":-49.25389,"address":"Av. do Cerrado, 999","tipo":"Capital"},
    {"name":"São Luís",            "state":"MA","lat":-2.52972, "lon":-44.30278,"address":"Rua Afonso Pena, s/n","tipo":"Capital"},
    {"name":"Cuiabá",              "state":"MT","lat":-15.59611,"lon":-56.09667,"address":"Av. General Mello, s/n","tipo":"Capital"},
    {"name":"Campo Grande",        "state":"MS","lat":-20.44278,"lon":-54.64611,"address":"Av. Afonso Pena, 3297","tipo":"Capital"},
    {"name":"Belo Horizonte",      "state":"MG","lat":-19.91722,"lon":-43.93444,"address":"Av. Afonso Pena, 1212","tipo":"Capital"},
    {"name":"Belém",               "state":"PA","lat":-1.45583, "lon":-48.50444,"address":"Praça Felipe Patroni, s/n","tipo":"Capital"},
    {"name":"João Pessoa",         "state":"PB","lat":-7.11528, "lon":-34.86278,"address":"Praça João Pessoa, s/n","tipo":"Capital"},
    {"name":"Curitiba",            "state":"PR","lat":-25.42944,"lon":-49.27167,"address":"Av. Cândido de Abreu, 817","tipo":"Capital"},
    {"name":"Recife",              "state":"PE","lat":-8.05361, "lon":-34.88111,"address":"Av. Cais do Apolo, 925","tipo":"Capital"},
    {"name":"Teresina",            "state":"PI","lat":-5.08917, "lon":-42.80194,"address":"R. Areolino de Abreu, 900","tipo":"Capital"},
    {"name":"Rio de Janeiro",      "state":"RJ","lat":-22.90278,"lon":-43.17444,"address":"R. Afonso Cavalcanti, 455","tipo":"Capital"},
    {"name":"Natal",               "state":"RN","lat":-5.79500, "lon":-35.21139,"address":"Av. Deodoro da Fonseca, 384","tipo":"Capital"},
    {"name":"Porto Velho",         "state":"RO","lat":-8.76194, "lon":-63.90389,"address":"Av. 7 de Setembro, 237","tipo":"Capital"},
    {"name":"Boa Vista",           "state":"RR","lat": 2.81972, "lon":-60.67333,"address":"Rua Coronel Pinto, 241","tipo":"Capital"},
    {"name":"Porto Alegre",        "state":"RS","lat":-30.03444,"lon":-51.21750,"address":"Av. Loureiro da Silva, 255","tipo":"Capital"},
    {"name":"Florianópolis",       "state":"SC","lat":-27.59500,"lon":-48.54861,"address":"R. Timóteo Pereira da Costa, 10","tipo":"Capital"},
    {"name":"Aracaju",             "state":"SE","lat":-10.91111,"lon":-37.07167,"address":"Av. Dr. Carlos Firpo, s/n","tipo":"Capital"},
    {"name":"São Paulo",           "state":"SP","lat":-23.55028,"lon":-46.63361,"address":"Viaduto do Chá, 15","tipo":"Capital"},
    {"name":"Palmas",              "state":"TO","lat":-10.18611,"lon":-48.33361,"address":"Quadra 502 Sul, Av. NS-02","tipo":"Capital"},
    # ── Interior (23 mais populosas não-capitais — Censo IBGE 2022) ───────────
    {"name":"Guarulhos",           "state":"SP","lat":-23.46278,"lon":-46.53333,"address":"Av. Bom Clima, 91 - Bom Clima","tipo":"Interior"},
    {"name":"Campinas",            "state":"SP","lat":-22.90556,"lon":-47.06083,"address":"Av. Anchieta, 200 - Centro","tipo":"Interior"},
    {"name":"São Gonçalo",         "state":"RJ","lat":-22.82694,"lon":-43.05417,"address":"R. Coronel Serrado, 530 - Centro","tipo":"Interior"},
    {"name":"São Bernardo do Campo","state":"SP","lat":-23.69417,"lon":-46.56472,"address":"R. Java, 425 - Jardim do Mar","tipo":"Interior"},
    {"name":"Duque de Caxias",     "state":"RJ","lat":-22.78500,"lon":-43.31167,"address":"Av. General Gurjão, s/n - Centro","tipo":"Interior"},
    {"name":"Nova Iguaçu",         "state":"RJ","lat":-22.75917,"lon":-43.45111,"address":"R. Dr. Moacyr Padilha, s/n - Centro","tipo":"Interior"},
    {"name":"Santo André",         "state":"SP","lat":-23.66444,"lon":-46.53278,"address":"Praça IV Centenário, s/n - Centro","tipo":"Interior"},
    {"name":"Osasco",              "state":"SP","lat":-23.53250,"lon":-46.79194,"address":"Av. Aryhovaldo de Miranda, 496 - Centro","tipo":"Interior"},
    {"name":"Sorocaba",            "state":"SP","lat":-23.50167,"lon":-47.45806,"address":"Av. Engenheiro Carlos Reinaldo Mendes, 3041","tipo":"Interior"},
    {"name":"Uberlândia",          "state":"MG","lat":-18.91861,"lon":-48.27722,"address":"Av. Anselmo Alves dos Santos, 600 - Santa Mônica","tipo":"Interior"},
    {"name":"Ribeirão Preto",      "state":"SP","lat":-21.17750,"lon":-47.81028,"address":"Praça das Bandeiras, 20 - Centro","tipo":"Interior"},
    {"name":"São José dos Campos", "state":"SP","lat":-23.17917,"lon":-45.88694,"address":"R. José de Alencar, 123 - Vila Santa Luzia","tipo":"Interior"},
    {"name":"Jaboatão dos Guararapes","state":"PE","lat":-8.11278,"lon":-35.01333,"address":"R. Padre Antônio Leite, s/n - Prazeres","tipo":"Interior"},
    {"name":"Contagem",            "state":"MG","lat":-19.93194,"lon":-44.05361,"address":"Av. João César de Oliveira, 3800 - Eldorado","tipo":"Interior"},
    {"name":"Joinville",           "state":"SC","lat":-26.30250,"lon":-48.84583,"address":"Praça Nereu Ramos, s/n - Centro","tipo":"Interior"},
    {"name":"Feira de Santana",    "state":"BA","lat":-12.26694,"lon":-38.96667,"address":"Praça da Prefeitura, s/n - Centro","tipo":"Interior"},
    {"name":"Londrina",            "state":"PR","lat":-23.31028,"lon":-51.15806,"address":"Av. Duque de Caxias, 635 - Centro","tipo":"Interior"},
    {"name":"Juiz de Fora",        "state":"MG","lat":-21.76194,"lon":-43.34972,"address":"Av. Brasil, 2001 - Bom Pastor","tipo":"Interior"},
    {"name":"Aparecida de Goiânia","state":"GO","lat":-16.82361,"lon":-49.24361,"address":"Av. das Nações, s/n - Cardoso","tipo":"Interior"},
    {"name":"Serra",               "state":"ES","lat":-20.12833,"lon":-40.30750,"address":"Av. Talma Rodrigues Ribeiro, 5416 - Portal de Jacaraípe","tipo":"Interior"},
    {"name":"Campos dos Goytacazes","state":"RJ","lat":-21.75500,"lon":-41.32472,"address":"Av. 28 de Março, s/n - Centro","tipo":"Interior"},
    {"name":"Ananindeua",          "state":"PA","lat":-1.36583, "lon":-48.37222,"address":"Rua 9 de Janeiro, s/n - Centro","tipo":"Interior"},
    {"name":"São José do Rio Preto","state":"SP","lat":-20.81694,"lon":-49.37583,"address":"Praça Rui Barbosa, s/n - Centro","tipo":"Interior"},
]

# ── Funções ────────────────────────────────────────────────────────────────────
def haversine(la1, lo1, la2, lo2):
    R = 6371
    dL, dO = math.radians(la2-la1), math.radians(lo2-lo1)
    a = math.sin(dL/2)**2 + math.cos(math.radians(la1))*math.cos(math.radians(la2))*math.sin(dO/2)**2
    return round(R*2*math.atan2(math.sqrt(a), math.sqrt(1-a)), 1)

def get_road_distance(c1, c2, ors_key):
    try:
        r = requests.post(
            "https://api.openrouteservice.org/v2/directions/driving-car",
            headers={"Authorization": ors_key, "Content-Type": "application/json"},
            json={"coordinates": [[c1["lon"],c1["lat"]], [c2["lon"],c2["lat"]]]},
            timeout=15
        )
        if r.status_code == 200:
            d = r.json()
            m = (d.get("routes",[{}])[0].get("summary",{}).get("distance") or
                 d["features"][0]["properties"]["segments"][0]["distance"])
            return round(m/1000, 1)
        return None
    except Exception:
        return None

def serp_idx(line, road):
    if line and road and line > 0:
        return round(road/line, 3)
    return None

# ── Excel ──────────────────────────────────────────────────────────────────────
def build_excel(cities, matrix):
    wb = Workbook()
    hf  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    hfd = PatternFill("solid", fgColor="0F1117")
    hfl = PatternFill("solid", fgColor="1A1D26")
    bf  = Font(name="Arial", bold=True, size=9)
    af  = PatternFill("solid", fgColor="F8FAFC")
    sf  = PatternFill("solid", fgColor="F1F5F9")
    thin = Side(style="thin", color="E2E8F0")
    bdr  = Border(bottom=thin, right=thin)

    def make_matrix(title, val_fn, font_fn):
        ws = wb.create_sheet(title)
        c0 = ws.cell(1,1,"Origem / Destino")
        c0.font = Font(name="Arial",bold=True,color="FFFFFF",size=8)
        c0.fill = hfl; c0.alignment = Alignment(horizontal="left",vertical="center")
        ws.column_dimensions["A"].width = 22
        for j,c in enumerate(cities,2):
            cell = ws.cell(1,j,f"{c['name']}, {c['state']}")
            cell.font = hf; cell.fill = hfd
            cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
            ws.column_dimensions[get_column_letter(j)].width = 14
        ws.row_dimensions[1].height = 28
        for i,c1 in enumerate(cities):
            r = i+2
            nc = ws.cell(r,1,f"{c1['name']}, {c1['state']}")
            nc.font = bf; nc.alignment = Alignment(vertical="center")
            if i%2==1: nc.fill = af
            for j,c2 in enumerate(cities,2):
                cell = ws.cell(r,j)
                cell.alignment = Alignment(horizontal="center",vertical="center")
                cell.border = bdr
                if c1["name"]==c2["name"]:
                    cell.value="—"; cell.fill=sf
                    cell.font=Font(name="Arial",color="94A3B8",size=10)
                else:
                    v = val_fn(c1,c2)
                    cell.value = v
                    cell.font = font_fn(v)
                    if isinstance(v,float): cell.number_format='#,##0.0'
                    if i%2==1 and v is not None: cell.fill=af
                ws.row_dimensions[r].height = 18
        return ws

    wb.remove(wb.active)

    # Aba 1 — Linha Reta
    make_matrix("Linha Reta (km)",
        lambda c1,c2: haversine(c1["lat"],c1["lon"],c2["lat"],c2["lon"]),
        lambda v: Font(name="Arial",bold=True,color="1E40AF",size=9) if isinstance(v,float)
                  else Font(name="Arial",size=9))

    # Aba 2 — Por Estrada
    def road_val(c1,c2):
        v = matrix.get(f"{c1['name']}-{c2['name']}",{}).get("road")
        return v
    def road_font(v):
        if isinstance(v,float): return Font(name="Arial",bold=True,color="166534",size=9)
        return Font(name="Arial",color="FCA5A5",size=8)
    ws2 = make_matrix("Por Estrada (km)", road_val, road_font)
    for row in ws2.iter_rows(min_row=2):
        for cell in row[1:]:
            if cell.value is None and cell.row>1 and cell.column>1:
                cell.value="N/D"

    # Aba 3 — Serpentividade
    serp_colors = {"Direta":("15803D","DCFCE7"), "Moderada":("92400E","FEF3C7"), "Sinuosa":("991B1B","FEE2E2")}
    def serp_val(c1,c2):
        d = matrix.get(f"{c1['name']}-{c2['name']}",{})
        return serp_idx(d.get("line"), d.get("road"))
    def serp_font_fn(v):
        if not isinstance(v,float): return Font(name="Arial",size=9)
        label,_,_ = serp_class(v)
        fc = serp_colors.get(label,("000000","FFFFFF"))[0]
        return Font(name="Arial",bold=True,color=fc,size=9)
    ws3 = make_matrix("Índice de Serpentividade", serp_val, serp_font_fn)
    # Colorir fundo das células
    for i,c1 in enumerate(cities):
        for j,c2 in enumerate(cities,2):
            if c1["name"]!=c2["name"]:
                d = matrix.get(f"{c1['name']}-{c2['name']}",{})
                v = serp_idx(d.get("line"),d.get("road"))
                if v:
                    label,_,_ = serp_class(v)
                    bg = serp_colors.get(label,("000000","FFFFFF"))[1]
                    ws3.cell(i+2,j).fill = PatternFill("solid",fgColor=bg)
                    ws3.cell(i+2,j).number_format='0.000'
                else:
                    ws3.cell(i+2,j).value="N/D"
                    ws3.cell(i+2,j).font=Font(name="Arial",color="FCA5A5",size=8)

    # Aba 4 — Ranking Serpentividade
    ws4 = wb.create_sheet("Ranking Serpentividade")
    headers4 = ["#","Cidade A","UF","Cidade B","UF","Reta (km)","Estrada (km)","Índice","Classificação"]
    widths4  = [4,20,5,20,5,12,14,10,14]
    for j,(h,w) in enumerate(zip(headers4,widths4),1):
        c=ws4.cell(1,j,h); c.font=hf; c.fill=hfd
        c.alignment=Alignment(horizontal="center",vertical="center")
        ws4.column_dimensions[get_column_letter(j)].width=w
    ws4.row_dimensions[1].height=22
    rows4=[]
    for c1,c2 in combinations(cities,2):
        d = matrix.get(f"{c1['name']}-{c2['name']}",{})
        line=d.get("line"); road=d.get("road")
        idx=serp_idx(line,road)
        if idx: rows4.append((c1,c2,line,road,idx))
    rows4.sort(key=lambda x: x[4], reverse=True)
    for rank,(c1,c2,line,road,idx) in enumerate(rows4,1):
        label,fc,bg=serp_class(idx)
        vals=[rank,c1["name"],c1["state"],c2["name"],c2["state"],line,road,idx,label]
        for j,v in enumerate(vals,1):
            cell=ws4.cell(rank+1,j,v)
            cell.alignment=Alignment(horizontal="center" if j not in [2,4] else "left",vertical="center")
            cell.border=bdr
            if rank%2==0: cell.fill=af
            if j==8:
                cell.font=Font(name="Arial",bold=True,color=fc[1:] if fc.startswith("#") else fc,size=9)
                cell.fill=PatternFill("solid",fgColor=bg[1:] if bg.startswith("#") else bg)
                cell.number_format='0.000'
            elif j==9:
                cell.font=Font(name="Arial",bold=True,color=fc[1:] if fc.startswith("#") else fc,size=9)
                cell.fill=PatternFill("solid",fgColor=bg[1:] if bg.startswith("#") else bg)
            elif j in [6,7]:
                cell.font=Font(name="Arial",size=9); cell.number_format='#,##0.0'
            else:
                cell.font=Font(name="Arial",size=9)
        ws4.row_dimensions[rank+1].height=16

    # Aba 5 — Coordenadas
    ws5=wb.create_sheet("Coordenadas das Sedes")
    h5=["Cidade","UF","Latitude","Longitude","Endereço da Sede"]
    w5=[20,5,14,14,44]
    for j,(h,w) in enumerate(zip(h5,w5),1):
        c=ws5.cell(1,j,h); c.font=hf; c.fill=hfd
        c.alignment=Alignment(horizontal="center",vertical="center")
        ws5.column_dimensions[get_column_letter(j)].width=w
    ws5.row_dimensions[1].height=22
    for i,c in enumerate(cities,2):
        for j,v in enumerate([c["name"],c["state"],c["lat"],c["lon"],c["address"]],1):
            cell=ws5.cell(i,j,v)
            cell.font=Font(name="Arial",size=9)
            cell.alignment=Alignment(horizontal="center" if j>2 else "left",vertical="center")
            cell.border=bdr
            if isinstance(v,float): cell.number_format='0.00000'
            if i%2==0: cell.fill=af
        ws5.row_dimensions[i].height=16

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ── Importar Excel existente ───────────────────────────────────────────────────
def import_from_xlsx(uploaded_file):
    """Lê o Excel exportado pelo app (qualquer versão) e retorna matrix + cidades."""
    import openpyxl
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)

    # Índice rápido de lookup: "Nome, UF" → dados completos do CAPITALS_DEFAULT
    default_index = {}
    for c in CAPITALS_DEFAULT:
        default_index[f"{c['name']}, {c['state']}"] = c
        default_index[c["name"]] = c  # fallback sem UF

    # ── Aba Coordenadas das Sedes (opcional — enriquece dados) ─────────────────
    coords_map = {}  # "Nome, UF" → {lat, lon, address, tipo}
    if "Coordenadas das Sedes" in wb.sheetnames:
        ws = wb["Coordenadas das Sedes"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            name, state = str(row[0]).strip(), str(row[1]).strip() if row[1] else ""
            key = f"{name}, {state}" if state else name
            try:
                lat = float(row[2]) if row[2] is not None else None
                lon = float(row[3]) if row[3] is not None else None
            except (TypeError, ValueError):
                lat, lon = None, None
            coords_map[key]  = {"lat": lat, "lon": lon,
                                "address": row[4] if len(row) > 4 else "",
                                "tipo": row[5] if len(row) > 5 else None}
            coords_map[name] = coords_map[key]  # fallback sem UF

    # ── Lê distâncias das abas ─────────────────────────────────────────────────
    def read_matrix_sheet(sheetname):
        data = {}
        if sheetname not in wb.sheetnames:
            return data
        ws = wb[sheetname]
        headers = [ws.cell(1, j).value for j in range(2, ws.max_column + 1)]
        for row in ws.iter_rows(min_row=2, min_col=1, values_only=True):
            c1 = row[0]
            if not c1: continue
            for j, val in enumerate(row[1:]):
                c2 = headers[j] if j < len(headers) else None
                if c2 and c1 != c2 and isinstance(val, (int, float)):
                    data[f"{c1}|{c2}"] = float(val)
        return data

    road_data = read_matrix_sheet("Por Estrada (km)")
    line_data = read_matrix_sheet("Linha Reta (km)")

    # ── Monta matrix interna ───────────────────────────────────────────────────
    matrix = {}
    all_pair_keys = set(road_data) | set(line_data)
    for pair_key in all_pair_keys:
        parts = pair_key.split("|")
        if len(parts) != 2: continue
        c1_full, c2_full = parts[0].strip(), parts[1].strip()
        c1_short = c1_full.split(",")[0].strip()
        c2_short = c2_full.split(",")[0].strip()
        road = road_data.get(pair_key)
        line = (line_data.get(pair_key) or
                line_data.get(f"{c2_full}|{c1_full}") or
                line_data.get(f"{c2_short}|{c1_short}"))
        entry = {"line": line, "road": road}
        matrix[f"{c1_short}-{c2_short}"] = entry
        matrix[f"{c2_short}-{c1_short}"] = entry

    # ── Descobre cidades únicas presentes no arquivo ───────────────────────────
    city_names_seen = {}  # nome_curto → nome_completo ("Nome, UF")
    for pair_key in all_pair_keys:
        for part in pair_key.split("|"):
            part = part.strip()
            short = part.split(",")[0].strip()
            if short not in city_names_seen:
                city_names_seen[short] = part

    # Monta lista de cidades: prioriza CAPITALS_DEFAULT, depois coords_map, depois mínimo
    imported_cities = []
    seen_names = set()
    # Primeiro passa pelas cidades do arquivo na ordem de CAPITALS_DEFAULT
    for c in CAPITALS_DEFAULT:
        full = f"{c['name']}, {c['state']}"
        if c["name"] in city_names_seen and c["name"] not in seen_names:
            imported_cities.append(dict(c))
            seen_names.add(c["name"])
    # Depois adiciona cidades do arquivo que não estão no CAPITALS_DEFAULT
    for short, full in city_names_seen.items():
        if short in seen_names:
            continue
        # Tenta enriquecer com coords_map ou default_index
        base = coords_map.get(full) or coords_map.get(short) or {}
        default = default_index.get(full) or default_index.get(short) or {}
        state = full.split(",")[1].strip() if "," in full else ""
        imported_cities.append({
            "name": short,
            "state": state or default.get("state", ""),
            "lat":  base.get("lat") or default.get("lat") or 0.0,
            "lon":  base.get("lon") or default.get("lon") or 0.0,
            "address": base.get("address") or default.get("address", ""),
            "tipo": base.get("tipo") or default.get("tipo", "Interior"),
        })
        seen_names.add(short)

    pairs_with_road = sum(1 for v in matrix.values() if v.get("road") is not None) // 2
    pairs_total = len(matrix) // 2

    return matrix, imported_cities, pairs_total, pairs_with_road

# ── Session state ──────────────────────────────────────────────────────────────
if "capitals" not in st.session_state:
    st.session_state.capitals = [dict(c) for c in CAPITALS_DEFAULT]
CAPITALS = st.session_state.capitals  # sempre aponta para a lista persistida

for k,v in [("selected",[c["name"] for c in CAPITALS]),("matrix",{}),
            ("calculated",False),("calc_cities",[]),("has_ors",False)]:
    if k not in st.session_state: st.session_state[k]=v

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuração")
    ors_key = st.text_input("Chave OpenRouteService", type="password",
        placeholder="5b3ce3597851...",
        help="Gratuita em openrouteservice.org — necessária para distâncias por estrada e índice de serpentividade")
    if ors_key:
        st.success("✓ Chave configurada")
    else:
        st.info("Sem chave: apenas linha reta ✈\n\n[Obter chave gratuita →](https://openrouteservice.org/dev/#/signup)")

    st.markdown("---")
    st.markdown("### 📍 Selecionar Cidades")
    c1b,c2b=st.columns(2)
    if c1b.button("✓ Todas",use_container_width=True):
        st.session_state.selected=[c["name"] for c in CAPITALS]
        for c in CAPITALS:
            st.session_state[f"cb_{c['name']}"]=True
        st.session_state.calculated=False; st.rerun()
    if c2b.button("✗ Nenhuma",use_container_width=True):
        st.session_state.selected=[]
        for c in CAPITALS:
            st.session_state[f"cb_{c['name']}"]=False
        st.session_state.calculated=False; st.rerun()
    new_sel=[]
    for c in CAPITALS:
        val = st.session_state.get(f"cb_{c['name']}", c["name"] in st.session_state.selected)
        if st.checkbox(f"{c['name']} ({c['state']})", value=val, key=f"cb_{c['name']}"):
            new_sel.append(c["name"])
    if set(new_sel)!=set(st.session_state.selected):
        st.session_state.selected=new_sel; st.session_state.calculated=False
    st.markdown(f"**{len(st.session_state.selected)}** cidade(s) selecionada(s)")
    st.markdown("---")
    st.markdown("### 📥 Importar Excel anterior")
    uploaded = st.file_uploader(
        "Carregue um .xlsx exportado pelo app",
        type=["xlsx"],
        help="Importa distâncias já calculadas — o app completará apenas os pares faltantes"
    )
    if uploaded:
        if st.button("⬆ Importar dados", use_container_width=True):
            try:
                import traceback
                matrix, imp_cities, total, with_road = import_from_xlsx(uploaded)

                # Atualiza CAPITALS com as cidades importadas que ainda não estejam na base
                existing_names = {c["name"] for c in st.session_state.capitals}
                for c in imp_cities:
                    if c["name"] not in existing_names:
                        st.session_state.capitals.append(dict(c))

                # Seleciona exatamente as cidades do arquivo (e reseta as outras)
                imp_names = [c["name"] for c in imp_cities]
                st.session_state.selected = imp_names
                # Força checkboxes
                for c in st.session_state.capitals:
                    st.session_state[f"cb_{c['name']}"] = c["name"] in imp_names

                st.session_state.matrix = matrix
                st.session_state.calc_cities = imp_cities
                st.session_state.calculated = True
                st.session_state.has_ors = with_road > 0

                # Pares faltantes (sem road) para poder retomar
                from itertools import combinations as comb2
                all_pairs = list(comb2(imp_cities, 2))
                pending = [(c1, c2) for c1, c2 in all_pairs
                           if matrix.get(f"{c1['name']}-{c2['name']}", {}).get("road") is None]
                n_done = len(all_pairs) - len(pending)
                st.session_state.pending_pairs = pending
                st.session_state.done_count = n_done
                st.session_state.total_pairs_count = len(all_pairs)

                st.success(f"✅ Importado! {with_road} pares com estrada · {len(pending)} pares faltando.")
                st.rerun()
            except Exception as e:
                st.error(f"Erro ao importar: {e}")
                st.code(traceback.format_exc())

# ── Main ───────────────────────────────────────────────────────────────────────
selected_cities=[c for c in CAPITALS if c["name"] in st.session_state.selected]

# Aviso se nenhuma cidade selecionada
if not selected_cities:
    st.warning("Selecione pelo menos 2 cidades no menu lateral.")

btn_col1, btn_col2, _ = st.columns([2,2,4])
with btn_col1:
    calc_clicked = st.button("⚡ Calcular Distâncias", disabled=len(selected_cities)<2,
                              use_container_width=True, type="primary")

# ── Cálculo ────────────────────────────────────────────────────────────────────
# ── Botão retomar ────────────────────────────────────────────────────────────
can_resume = (st.session_state.get("pending_pairs") and
              st.session_state.get("calc_cities") == selected_cities)

with btn_col1:
    pass  # já renderizado acima

resume_clicked = False
if can_resume:
    pending = st.session_state.pending_pairs
    done_count = st.session_state.get("done_count", 0)
    total_count = st.session_state.get("total_pairs_count", 0)
    st.info(f"⚠️ Cálculo incompleto: **{done_count}/{total_count}** pares concluídos. "
            f"Restam **{len(pending)}** pares.")
    resume_clicked = st.button("▶ Retomar Cálculo", use_container_width=False, type="secondary")

def run_calculation(pairs_to_calc, ors_key, has_ors, existing_matrix):
    matrix = dict(existing_matrix)
    all_pairs_count = st.session_state.get("total_pairs_count", len(pairs_to_calc))
    done_so_far = st.session_state.get("done_count", 0)

    prog = st.progress(int(done_so_far/all_pairs_count*100), text="Iniciando…")
    status = st.empty()
    remaining = list(pairs_to_calc)

    for idx, (c1, c2) in enumerate(remaining):
        line = haversine(c1["lat"], c1["lon"], c2["lat"], c2["lon"])
        road = None
        if has_ors:
            status.markdown(f"🚗 **{c1['name']} → {c2['name']}** "
                            f"({done_so_far+idx+1}/{all_pairs_count})")
            road = get_road_distance(c1, c2, ors_key)

        key = f"{c1['name']}-{c2['name']}"
        matrix[key] = {"line": line, "road": road}
        matrix[f"{c2['name']}-{c1['name']}"] = {"line": line, "road": road}

        # Salva progresso parcial a cada par
        st.session_state.matrix = matrix
        st.session_state.done_count = done_so_far + idx + 1
        st.session_state.pending_pairs = remaining[idx+1:]

        pct = int((done_so_far+idx+1)/all_pairs_count*100)
        prog.progress(pct, text=f"Calculando… {done_so_far+idx+1}/{all_pairs_count}")

    prog.progress(100, text="✅ Concluído!")
    status.empty()
    st.session_state.pending_pairs = []
    return matrix

if calc_clicked and len(selected_cities)>=2:
    pairs=list(combinations(selected_cities,2))
    has_ors=bool(ors_key and len(ors_key)>10)

    if has_ors:
        test=get_road_distance(selected_cities[0],selected_cities[1],ors_key)
        if test is None:
            st.error("❌ Chave ORS inválida ou sem créditos. Calculando apenas linha reta.")
            has_ors=False

    # Inicializa estado
    st.session_state.pending_pairs = pairs
    st.session_state.done_count = 0
    st.session_state.total_pairs_count = len(pairs)
    st.session_state.calc_cities = selected_cities
    st.session_state.has_ors = has_ors
    st.session_state.calculated = False

    matrix = run_calculation(pairs, ors_key, has_ors, {})
    st.session_state.matrix = matrix
    st.session_state.calculated = True
    st.session_state.has_ors = has_ors

elif resume_clicked:
    has_ors = st.session_state.get("has_ors", False)
    pending = st.session_state.pending_pairs
    existing = st.session_state.get("matrix", {})

    if has_ors and not ors_key:
        st.error("Cole a chave ORS novamente para retomar.")
    else:
        matrix = run_calculation(pending, ors_key, has_ors, existing)
        st.session_state.matrix = matrix
        st.session_state.calculated = True

# ── Resultados ─────────────────────────────────────────────────────────────────

# Aba de Cidades sempre visível (mesmo antes de calcular)
st.markdown("---")

# ── TABS PRINCIPAIS ────────────────────────────────────────────────────────────
tab_calc, tab_serp, tab_cities = st.tabs([
    "📊 Matriz de Distâncias",
    "🐍 Índice de Serpentividade",
    "📍 Cidades & Geolocalização"
])

# ── TAB CIDADES ────────────────────────────────────────────────────────────────
with tab_cities:
    st.markdown("#### Base de Cidades")
    st.markdown("Cidades disponíveis para cálculo. Selecione as desejadas no menu lateral.")

    # Tabela de cidades com indicação de selecionada
    tbl_c = ("<div style='overflow-x:auto'><table class='dist-table'>"
             "<thead><tr>"
             "<th style='text-align:left'>Cidade</th>"
             "<th>UF</th>"
             "<th>Tipo</th>"
             "<th>Latitude (Sede)</th>"
             "<th>Longitude (Sede)</th>"
             "<th>Endereço da Sede</th>"
             "<th>Selecionada</th>"
             "</tr></thead><tbody>")
    for i, c in enumerate(CAPITALS):
        sel = c["name"] in st.session_state.selected
        tipo = c.get("tipo", "—")
        tipo_badge = (
            "<span style='background:#eff6ff;color:#1e40af;padding:2px 8px;border-radius:8px;font-size:.68rem;font-weight:700'>Capital</span>"
            if tipo == "Capital" else
            "<span style='background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:8px;font-size:.68rem;font-weight:700'>Interior</span>"
        )
        sel_badge = ("<span style='background:#dcfce7;color:#15803d;padding:2px 9px;"
                 "border-radius:10px;font-size:.72rem;font-weight:700'>✓ Sim</span>" if sel else
                 "<span style='background:#f1f5f9;color:#94a3b8;padding:2px 9px;"
                 "border-radius:10px;font-size:.72rem'>— Não</span>")
        bg = "background:#f0fdf4;" if sel else ""
        tbl_c += (f"<tr style='{bg}'>"
                  f"<td style='font-weight:600;text-align:left'>{c['name']}</td>"
                  f"<td style='text-align:center;font-family:monospace;font-size:.8rem'>{c['state']}</td>"
                  f"<td style='text-align:center'>{tipo_badge}</td>"
                  f"<td style='text-align:center;font-family:monospace;font-size:.78rem'>{c.get('lat',0):.5f}</td>"
                  f"<td style='text-align:center;font-family:monospace;font-size:.78rem'>{c.get('lon',0):.5f}</td>"
                  f"<td style='text-align:left;font-size:.78rem;color:#64748b'>{c.get('address','')}</td>"
                  f"<td style='text-align:center'>{sel_badge}</td>"
                  f"</tr>")
    tbl_c += "</tbody></table></div>"
    st.markdown(tbl_c, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### ➕ Adicionar Nova Cidade")
    st.caption("Preencha os dados da sede municipal (prefeitura) da cidade que deseja adicionar.")

    with st.form("form_add_city", clear_on_submit=True):
        fc1, fc2, fc3 = st.columns([3,1,1])
        fc4, fc5 = st.columns(2)
        fc6, fc7 = st.columns([3,1])
        new_name    = fc1.text_input("Nome da Cidade", placeholder="Ex: Campinas")
        new_state   = fc2.text_input("UF", placeholder="SP", max_chars=2)
        new_tipo    = fc3.selectbox("Tipo", ["Interior", "Capital"])
        new_lat     = fc4.number_input("Latitude (Sede)", value=-23.0, format="%.5f", step=0.00001)
        new_lon     = fc5.number_input("Longitude (Sede)", value=-47.0, format="%.5f", step=0.00001)
        new_address = fc6.text_input("Endereço da Sede", placeholder="Rua XV de Novembro, 1000 - Centro")
        submitted   = fc7.form_submit_button("➕ Adicionar", use_container_width=True, type="primary")

        if submitted:
            if not new_name or not new_state:
                st.error("Nome e UF são obrigatórios.")
            elif any(c["name"].lower() == new_name.strip().lower() for c in st.session_state.capitals):
                st.warning(f"'{new_name}' já está na base.")
            else:
                new_city = {
                    "name": new_name.strip(),
                    "state": new_state.strip().upper(),
                    "lat": float(new_lat),
                    "lon": float(new_lon),
                    "address": new_address.strip(),
                    "tipo": new_tipo
                }
                st.session_state.capitals.append(new_city)
                st.session_state.selected.append(new_name.strip())
                st.session_state[f"cb_{new_name.strip()}"] = True
                st.success(f"✅ '{new_name.strip()}, {new_state.strip().upper()}' adicionada!")
                st.rerun()

if st.session_state.calculated and st.session_state.calc_cities:
    cities=st.session_state.calc_cities
    matrix=st.session_state.matrix
    has_ors=st.session_state.has_ors

    if not has_ors:
        st.info("ℹ️ Distâncias por estrada e índice de serpentividade indisponíveis sem a chave ORS.")

    # ── Tabs de resultados ─────────────────────────────────────────────────────
    tab1, tab2 = tab_calc, tab_serp  # alias para os resultados

    # ── TAB 1: Matriz ──────────────────────────────────────────────────────────
    with tab1:
        st.markdown("""
        <div style='display:flex;gap:20px;margin-bottom:12px;font-size:.78rem;color:#64748b;font-family:monospace'>
            <span><b style='color:#1e40af'>■</b> ✈ Linha reta (km)</span>
            <span><b style='color:#166534'>■</b> 🚗 Por estrada (km)</span>
            <span><b style='color:#15803d'>■</b> 🐍 Índice serpentividade</span>
        </div>""", unsafe_allow_html=True)

        tbl="<div style='overflow-x:auto'><table class='dist-table'><thead><tr><th>Origem / Destino</th>"
        for c in cities:
            tbl+=f"<th>{c['name']}<br><span style='font-family:monospace;font-size:.58rem;opacity:.6'>{c['state']}</span></th>"
        tbl+="</tr></thead><tbody>"
        for c1 in cities:
            tbl+=f"<tr><td>{c1['name']}, {c1['state']}</td>"
            for c2 in cities:
                if c1["name"]==c2["name"]:
                    tbl+="<td class='cell-self'>—</td>"
                else:
                    d=matrix.get(f"{c1['name']}-{c2['name']}",{})
                    line=d.get("line"); road=d.get("road")
                    idx=serp_idx(line,road)
                    road_html=(f"<div><span class='val-road'>{road} km</span> <span class='val-label'>🚗</span></div>"
                               if road else "<div><span class='val-nd'>N/D</span> <span class='val-label'>🚗</span></div>")
                    if idx:
                        label,fc,bg=serp_class(idx)
                        serp_html=(f"<div><span class='serp-badge' style='color:{fc};background:{bg}'>"
                                   f"{idx:.3f} · {label}</span></div>")
                    else:
                        serp_html=""
                    tbl+=(f"<td><div><span class='val-line'>{line} km</span> <span class='val-label'>✈</span></div>"
                          f"{road_html}{serp_html}</td>")
            tbl+="</tr>"
        tbl+="</tbody></table></div>"
        st.markdown(tbl, unsafe_allow_html=True)

    # ── TAB 2: Ranking Serpentividade ──────────────────────────────────────────
    with tab2:
        st.markdown("""
        <div style='margin-bottom:14px;font-size:.82rem;color:#475569;line-height:1.6'>
            O <b>Índice de Serpentividade</b> = Distância por Estrada ÷ Distância em Linha Reta.<br>
            Quanto mais próximo de <b>1.0</b>, mais direta é a estrada. Valores altos indicam rotas sinuosas.
        </div>
        <div style='display:flex;gap:12px;margin-bottom:16px;flex-wrap:wrap'>
            <span style='background:#dcfce7;color:#15803d;padding:4px 12px;border-radius:10px;font-size:.78rem;font-family:monospace;font-weight:700'>🟢 &lt; 1.20 — Direta</span>
            <span style='background:#fef3c7;color:#92400e;padding:4px 12px;border-radius:10px;font-size:.78rem;font-family:monospace;font-weight:700'>🟡 1.20–1.50 — Moderada</span>
            <span style='background:#fee2e2;color:#991b1b;padding:4px 12px;border-radius:10px;font-size:.78rem;font-family:monospace;font-weight:700'>🔴 &gt; 1.50 — Sinuosa</span>
        </div>
        """, unsafe_allow_html=True)

        if not has_ors:
            st.warning("Configure a chave ORS para calcular o índice de serpentividade.")
        else:
            # Monta todos os pares com índice calculado
            all_rows=[]
            for c1,c2 in combinations(cities,2):
                d=matrix.get(f"{c1['name']}-{c2['name']}", {})
                line=d.get("line"); road=d.get("road")
                idx=serp_idx(line,road)
                if idx: all_rows.append({"c1":c1,"c2":c2,"line":line,"road":road,"idx":idx})
            all_rows.sort(key=lambda x:x["idx"],reverse=True)

            # ── Filtros ────────────────────────────────────────────────────────
            st.markdown("##### Filtrar ranking")
            f_col1, f_col2 = st.columns([2,3])
            filter_tipo = f_col1.radio(
                "Tipo de cidades",
                ["Todas", "Apenas Capitais", "Apenas Interior", "Mesmo Estado"],
                horizontal=False, label_visibility="collapsed"
            )
            from collections import Counter
            uf_counts = Counter(c["state"] for c in cities)
            ufs_multi = sorted([uf for uf,cnt in uf_counts.items() if cnt > 1])
            uf_filter = None
            if filter_tipo == "Mesmo Estado":
                if ufs_multi:
                    uf_filter = f_col2.selectbox("Estado (UF)", ufs_multi)
                else:
                    f_col2.warning("Nenhum estado tem mais de uma cidade selecionada.")

            def match_filter(r):
                t1=r["c1"].get("tipo","—"); t2=r["c2"].get("tipo","—")
                s1=r["c1"]["state"];        s2=r["c2"]["state"]
                if filter_tipo=="Apenas Capitais":  return t1=="Capital" and t2=="Capital"
                if filter_tipo=="Apenas Interior":  return t1=="Interior" and t2=="Interior"
                if filter_tipo=="Mesmo Estado":     return bool(uf_filter and s1==uf_filter and s2==uf_filter)
                return True

            rows=[r for r in all_rows if match_filter(r)]

            if not rows:
                st.info("Nenhum par encontrado. Selecione mais cidades ou mude o filtro.")
            else:
                def tipo_chip(c):
                    t=c.get("tipo","—")
                    col,bg=("#1e40af","#eff6ff") if t=="Capital" else ("#92400e","#fef3c7")
                    return f"<span style='background:{bg};color:{col};padding:1px 6px;border-radius:6px;font-size:.65rem;font-weight:700'>{t}</span>"

                tbl2=("<div style='overflow-x:auto'><table class='serp-table'>"
                      "<thead><tr><th>#</th><th>Cidade A</th><th>UF</th><th>Tipo</th>"
                      "<th>Cidade B</th><th>UF</th><th>Tipo</th>"
                      "<th>Reta (km)</th><th>Estrada (km)</th><th>Índice</th><th>Classif.</th>"
                      "</tr></thead><tbody>")
                for rank,r in enumerate(rows,1):
                    label,fc,bg=serp_class(r["idx"])
                    badge=f"<span class='serp-badge' style='color:{fc};background:{bg}'>{label}</span>"
                    tbl2+=(f"<tr>"
                           f"<td style='text-align:center;color:#94a3b8;font-size:.75rem'>{rank}</td>"
                           f"<td><b>{r['c1']['name']}</b></td>"
                           f"<td style='text-align:center;font-family:monospace;font-size:.75rem'>{r['c1']['state']}</td>"
                           f"<td>{tipo_chip(r['c1'])}</td>"
                           f"<td><b>{r['c2']['name']}</b></td>"
                           f"<td style='text-align:center;font-family:monospace;font-size:.75rem'>{r['c2']['state']}</td>"
                           f"<td>{tipo_chip(r['c2'])}</td>"
                           f"<td style='text-align:center;color:#1e40af;font-weight:700'>{r['line']}</td>"
                           f"<td style='text-align:center;color:#166534;font-weight:700'>{r['road']}</td>"
                           f"<td style='text-align:center'><span style='font-family:monospace;font-weight:700;color:{fc}'>{r['idx']:.3f}</span></td>"
                           f"<td style='text-align:center'>{badge}</td>"
                           f"</tr>")
                tbl2+="</tbody></table></div>"
                st.markdown(tbl2, unsafe_allow_html=True)

                st.markdown("---")
                idxs=[r["idx"] for r in rows]
                m1,m2,m3,m4=st.columns(4)
                m1.metric("Média geral", f"{sum(idxs)/len(idxs):.3f}")
                m2.metric("Mais direta 🟢", f"{min(idxs):.3f}",
                          f"{rows[-1]['c1']['name']} ↔ {rows[-1]['c2']['name']}")
                m3.metric("Mais sinuosa 🔴", f"{max(idxs):.3f}",
                          f"{rows[0]['c1']['name']} ↔ {rows[0]['c2']['name']}")
                diretas=sum(1 for i in idxs if i<1.2)
                m4.metric("Rotas diretas (< 1.20)", f"{diretas}/{len(idxs)}")

    # ── Download ───────────────────────────────────────────────────────────────
    st.markdown("---")
    with btn_col2:
        excel=build_excel(cities,matrix)
        st.download_button("📊 Baixar Excel", data=excel,
            file_name="distancias-cidades.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
